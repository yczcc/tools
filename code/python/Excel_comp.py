#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os, sys
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter

reload(sys)
sys.setdefaultencoding('utf-8')

SOURCE_FILE1="C:\Users\changczhou\Desktop\https.xlsx"

MIG_OK_CONFIG_DOMAIN = []
MIG_OK_WHITE_DOMAIN = []
MIG_NOT_DOMAIN = []

widths = [
    (126,    1), (159,    0), (687,     1), (710,   0), (711,   1),
    (727,    0), (733,    1), (879,     0), (1154,  1), (1161,  0),
    (4347,   1), (4447,   2), (7467,    1), (7521,  0), (8369,  1),
    (8426,   0), (9000,   1), (9002,    2), (11021, 1), (12350, 2),
    (12351,  1), (12438,  2), (12442,   0), (19893, 2), (19967, 1),
    (55203,  2), (63743,  1), (64106,   2), (65039, 1), (65059, 0),
    (65131,  2), (65279,  1), (65376,   2), (65500, 1), (65510, 2),
    (120831, 1), (262141, 2), (1114109, 1),
]

def get_width( o ):
    """Return the screen column width for unicode ordinal o."""
    global widths
    if o == 0xe or o == 0xf:
        return 0
    for num, wid in widths:
        if o <= num:
            return wid
    return 1

def chr_width(str):
    len = 0
    for chr in str:
        len = len + get_width(chr)
    return len

def classify_mig_security_https_status(filePath):
    MIG_OK_CONFIG_DOMAIN[:] = []
    MIG_OK_WHITE_DOMAIN[:] = []
    MIG_NOT_DOMAIN[:] = []
    if os.path.exists(filePath) == False:
        print "[ERROR] excel file:", filePath, "do not exit"
        return 1
    wb = load_workbook(filePath)
    listSheet = wb.worksheets
    listSheetLen = len(listSheet)
    if 1 > listSheetLen:
        print "[ERROR] excel file data error, no sheet"
        return 2

    sheet0 = wb.worksheets[0]
    if "切换HTTPS进展-按域名维度(绿色代表已完成)" != sheet0.title:
        print "[ERROR] excel file data error, first sheet:", sheet0.title
        return 3

    if 1 < listSheetLen:
        for index in range(1, listSheetLen):
            wb.remove(wb.worksheets[1])

    # 创建分类的sheet
    sheet1 = wb.create_sheet(unicode("已跳转或仅HTTPS", "utf-8"), index=1)
    sheet2 = wb.create_sheet(unicode("已加白名单", "utf-8"), index=2)
    sheet3 = wb.create_sheet(unicode("未达标", "utf-8"), index=3)

    # 进行数据分类
    sheet0RowNum = sheet0.max_row
    sheet0ColumnNum = sheet0.max_column
    if 11 != sheet0ColumnNum:
        print "[ERROR] excel file data max_column error, first sheet:", sheet0.title, " max_column:", sheet0ColumnNum
        return
    sheet0RowIndex = 0
    sheet1RowNum = 0
    sheet2RowNum = 0
    sheet3RowNum = 0
    rows = sheet0.rows
    for row in rows:
        sheet0RowIndex = sheet0RowIndex + 1
        if 1 == sheet0RowIndex:
            sheet1.append([col.value for col in row])
            sheet2.append([col.value for col in row])
            sheet3.append([col.value for col in row])
            bold_font = Font(bold=True)
            for i in range(1, sheet0ColumnNum + 1):
                sheet1.cell(row=1, column=i).font = bold_font
                sheet2.cell(row=1, column=i).font = bold_font
                sheet3.cell(row=1, column=i).font = bold_font
            continue

        sheet0Cell = sheet0.cell(row=sheet0RowIndex, column=10)
        if '1' == sheet0Cell.value:
            sheet1.append([col.value for col in row])
            sheet1RowNum = sheet1RowNum + 1
            MIG_OK_CONFIG_DOMAIN.append(sheet0.cell(row=sheet0RowIndex, column=4))
            continue

        sheet0Cell = sheet0.cell(row=sheet0RowIndex, column=11)
        if '1' == sheet0Cell.value:
            sheet2.append([col.value for col in row])
            sheet2RowNum = sheet2RowNum + 1
            MIG_OK_WHITE_DOMAIN.append(sheet0.cell(row=sheet0RowIndex, column=4))
            continue

        sheet3.append([col.value for col in row])
        sheet3RowNum = sheet3RowNum + 1
        MIG_NOT_DOMAIN.append(sheet0.cell(row=sheet0RowIndex, column=4))
    print "got total:", sheet0RowNum-1, " rows, sheet1 total:", sheet1RowNum, \
        " rows, sheet2 total:", sheet2RowNum, " rows, other total:", sheet3RowNum," rows."
    wb.save(filePath)
    wb.close()
    return 0

def manage_https(sa_mig_https_file, sa_teg_https_file):
    if os.path.exists(sa_mig_https_file) == False:
        print "[ERROR] excel file:", sa_mig_https_file, "do not exit"
        return 1
    wbMigHttps = load_workbook(sa_mig_https_file)

    if os.path.exists(sa_teg_https_file) == False:
        print "[ERROR] excel file:", sa_teg_https_file, "do not exit"
        return 2
    wbTegHttps = load_workbook(sa_teg_https_file)

    MIG_OK_CONFIG_DOMAIN[:] = []
    MIG_OK_WHITE_DOMAIN[:] = []
    MIG_NOT_DOMAIN[:] = []
    # 获取当前mig分类好的数据
    listMigHttpsSheet = wbMigHttps.worksheets
    listMigHttpsSheetLen = len(listMigHttpsSheet)
    if 4 != listMigHttpsSheetLen:
        print "[ERROR] excel file data error, listMigHttpsSheetLen:",listMigHttpsSheetLen
        return 2
    sheetMigHttps1 = wbMigHttps.worksheets[1]
    if "已跳转或仅HTTPS" != sheetMigHttps1.title:
        print "[ERROR] excel file data error, first sheet:", sheetMigHttps1.title
        return 3
    sheetMigHttps2 = wbMigHttps.worksheets[2]
    if "已加白名单" != sheetMigHttps2.title:
        print "[ERROR] excel file data error, first sheet:", sheetMigHttps2.title
        return 3
    sheetMigHttps3 = wbMigHttps.worksheets[3]
    if "未达标" != sheetMigHttps3.title:
        print "[ERROR] excel file data error, first sheet:", sheetMigHttps3.title
        return 3
    for i in range(2, sheetMigHttps1.max_row):
        MIG_OK_CONFIG_DOMAIN.append(sheetMigHttps1.cell(row=i, column=4).value)
    for i in range(2, sheetMigHttps2.max_row):
        MIG_OK_WHITE_DOMAIN.append(sheetMigHttps2.cell(row=i, column=4).value)
    for i in range(2, sheetMigHttps3.max_row):
        MIG_NOT_DOMAIN.append(sheetMigHttps3.cell(row=i, column=4).value)
    print "got ok_config:", len(MIG_OK_CONFIG_DOMAIN), " ok_white:", len(MIG_OK_WHITE_DOMAIN),\
        " not_ok:", len(MIG_NOT_DOMAIN)

    listTegHttpsSheet = wbTegHttps.worksheets
    listTegHttpsSheetLen = len(listTegHttpsSheet)
    if 1 > listTegHttpsSheetLen:
        print "[ERROR] excel file data error, no sheet"
        return 2

    sheet0 = wbTegHttps.worksheets[0]
    # if "Worksheet" != sheet0.title:
    #     print "[ERROR] excel file data error, first sheet:", sheet0.title
    #     return 3

    # if 1 < listTegHttpsSheetLen:
    #     for index in range(1, listTegHttpsSheetLen):
    #         wbTegHttps.remove(wbTegHttps.worksheets[1])

    sheet0ColumnNum = sheet0.max_column
    if 9 > sheet0ColumnNum:
        print "[ERROR] excel file data error, sheet0ColumnNum:", sheet0ColumnNum
        return 3
    elif 9 < sheet0ColumnNum:
        sheet0.delete_cols(10, sheet0ColumnNum - 9)
        sheet0ColumnNum = sheet0.max_column

    sheet0RowIndex = 0
    sheet0ColumnWidths = [0, 0, 0, 0, 0]
    rows = sheet0.rows
    for row in rows:
        sheet0RowIndex = sheet0RowIndex + 1
        if 1 == sheet0RowIndex:
            cell = sheet0.cell(row=sheet0RowIndex, column=2)
            if "域名" != cell.value:
                print "[ERROR] excel file data error, first row:", cell.value
                return 4
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 1).value = "MIG达标-配置"
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 2).value = "MIG达标-白名单"
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 3).value = "MIG未达标"
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 4).value = "MIG未知"
            # 计算列宽，用于自动调整
            for i in range(0, 5):
                cellValue = sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum+i).value
                cellLen = chr_width(cellValue.encode("gbk"))
                if cellLen > sheet0ColumnWidths[i]:
                    sheet0ColumnWidths[i] = cellLen
            continue

        sheet0Cell = sheet0.cell(row=sheet0RowIndex, column=2)
        if sheet0Cell.value in MIG_OK_CONFIG_DOMAIN:
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 1).value = '1'
        elif sheet0Cell.value in MIG_OK_WHITE_DOMAIN:
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 2).value = '1'
        elif sheet0Cell.value in MIG_NOT_DOMAIN:
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 3).value = '1'
        else:
            sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + 4).value = '1'

        # 计算列宽，用于自动调整
        for i in range(0, 5):
            cellValue = sheet0.cell(row=sheet0RowIndex, column=sheet0ColumnNum + i).value
            if cellValue is None:
                continue
            cellLen = chr_width(cellValue.encode("gbk"))
            if cellLen > sheet0ColumnWidths[i]:
                sheet0ColumnWidths[i] = cellLen
    # 调整列宽
    for i, column_width in enumerate(sheet0ColumnWidths):
        print i, get_column_letter(sheet0ColumnNum + i), column_width
        sheet0.column_dimensions[get_column_letter(sheet0ColumnNum + i)].width = column_width
    wbTegHttps.save(sa_teg_https_file)
    wbTegHttps.close()
    return 0

sa_mig_security_https = "C:\Users\changczhou\Desktop\db.t_md_sa_mig_security_https_0.xlsx"
sa_teg_https_file = "C:\Users\changczhou\Desktop\HTTPS_2019-03-01.xlsx"

if __name__ == '__main__':
    # if len(sys.argv) == 1:
    #     raw_input("[error] need root dir")
    #     sys.exit(-1)
    # convertdir = sys.argv[1]
    # filePath1 = SOURCE_FILE1
    # read_excel(filePath1)

    # ret = classify_mig_security_https_status(sa_mig_security_https)
    # if 0 != ret:
    #     print "[ERROR] classify_mig_security_https_status failed, ret:", ret
    # else:
    #     print "[SUCCESS] classify_mig_security_https_status done"

    ret = manage_https(sa_mig_security_https, sa_teg_https_file)
    if 0 != ret:
        print "[ERROR] manage_https failed, ret:", ret
    else:
        print "[SUCCESS] manage_https done"

